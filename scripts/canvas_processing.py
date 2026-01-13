"""
Common processing utilities for Canvas uploads.
Handles markdown conversion, LaTeX math processing, image uploads, and code highlighting.
"""

import re
import os
import requests
import markdown
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import base64
from urllib.parse import urljoin

try:
    from openpyxl import load_workbook
    from openpyxl.styles.colors import COLOR_INDEX
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def convert_align_to_array(content: str) -> str:
    """Convert LaTeX align environments to Canvas-compatible array format."""
    content = re.sub(r'\\begin\{align\*?\}', r'\\begin{array}{rl}', content)
    content = re.sub(r'\\end\{align\*?\}', r'\\end{array}', content)
    return content

def process_math_block(content: str) -> str:
    """Process math blocks to handle align environments and wrapping."""
    def replace_math_block(match):
        math_content = match.group(1)
        math_content = convert_align_to_array(math_content)
        return f'$${math_content}$$'

    # Handle standalone align environments (wrap in $$)
    content = re.sub(r'(?<!\$\$)\s*\\begin\{align\*?\}(.*?)\\end\{align\*?\}\s*(?!\$\$)',
                     replace_math_block, content, flags=re.DOTALL)

    return content

def wrap_standalone_align(content: str) -> str:
    """Wrap standalone align environments in $$ and convert to array."""
    def replace_align(match):
        align_content = match.group(1)
        array_content = convert_align_to_array(f'\\begin{{align}}{align_content}\\end{{align}}')
        return f'$${array_content}$$'

    # Match standalone align environments (not already wrapped in $$)
    pattern = r'(?<!\$\$)\s*\\begin\{align\*?\}(.*?)\\end\{align\*?\}\s*(?!\$\$)'
    content = re.sub(pattern, replace_align, content, flags=re.DOTALL)

    return content

def protect_math_content(content: str) -> Tuple[str, Dict[str, str]]:
    """Protect math content from markdown processing by replacing with placeholders."""
    math_blocks = {}
    counter = 0

    def protect_display_math(match):
        nonlocal counter
        math_content = match.group(1)
        # Convert align to array BEFORE protection
        math_content = convert_align_to_array(math_content)
        # Double all backslashes to preserve them through markdown processing
        math_content = math_content.replace('\\', '\\\\')
        # Protect underscores, asterisks, and tildes from being treated as markdown
        math_content = math_content.replace('_', '§UNDERSCORE§')
        math_content = math_content.replace('*', '§ASTERISK§')
        math_content = math_content.replace('~', '§TILDE§')
        placeholder = f'§MATH_DISPLAY_START§{math_content}§MATH_DISPLAY_END§'
        return placeholder

    def protect_inline_math(match):
        nonlocal counter
        math_content = match.group(1)
        # Double all backslashes to preserve them through markdown processing
        math_content = math_content.replace('\\', '\\\\')
        # Protect underscores, asterisks, and tildes from being treated as markdown
        math_content = math_content.replace('_', '§UNDERSCORE§')
        math_content = math_content.replace('*', '§ASTERISK§')
        math_content = math_content.replace('~', '§TILDE§')
        placeholder = f'§MATH_INLINE_START§{math_content}§MATH_INLINE_END§'
        return placeholder

    # First handle $$...$$ (including environments) - convert align to array
    content = re.sub(r'\$\$(.+?)\$\$', protect_display_math, content, flags=re.DOTALL)

    # Then handle $...$ for inline (but not $$) - convert to \(...\)
    content = re.sub(r'(?<!\$)\$((?:[^$]|\\\$)+?)\$(?!\$)', protect_inline_math, content)

    return content, math_blocks

def restore_math_content(content: str, math_blocks: Dict[str, str]) -> str:
    """Restore protected math content with Canvas-compatible delimiters."""
    # Restore math delimiters - Canvas uses \(...\) for inline and $$...$$ for display
    content = content.replace('§MATH_INLINE_START§', r'\(')
    content = content.replace('§MATH_INLINE_END§', r'\)')
    content = content.replace('§MATH_DISPLAY_START§', '$$')
    content = content.replace('§MATH_DISPLAY_END§', '$$')

    # Fix double-escaped backslashes and restore protected characters in math
    def fix_math_backslashes(match):
        math_content = match.group(0)
        # Undo the backslash doubling we did during protection
        # We doubled all \ to \\ to protect them through markdown
        # Now we need to convert them back: \\\\ -> \\
        math_content = math_content.replace('\\\\', '\\')
        # Restore special LaTeX line breaks as double backslashes for MathJax
        # Use regex to handle any spacing variations around the token
        math_content = re.sub(r'\s*§LATEXLINEBREAK§\s*', r' \\\\ ', math_content)
        # Restore asterisks, underscores, and tildes that we protected
        math_content = math_content.replace('§ASTERISK§', '*')
        math_content = math_content.replace('§UNDERSCORE§', '_')
        math_content = math_content.replace('§TILDE§', '~')
        return math_content

    # Fix in both display ($$...$$) and inline (\(...\)) math
    content = re.sub(r'\$\$.*?\$\$', fix_math_backslashes, content, flags=re.DOTALL)
    content = re.sub(r'\\\(.*?\\\)', fix_math_backslashes, content, flags=re.DOTALL)

    return content

def preserve_linebreaks_in_math(content: str) -> str:
    """Replace \\\\ in math environments with special tokens to preserve through processing."""
    def replace_linebreaks(match):
        math_content = match.group(1)
        # Replace double backslashes with special token
        math_content = math_content.replace('\\\\', '§LATEXLINEBREAK§')
        return f'$${math_content}$$'

    # Process display math blocks first
    content = re.sub(r'\$\$(.*?)\$\$', replace_linebreaks, content, flags=re.DOTALL)

    return content

def upload_image_to_canvas(image_path: str, canvas_base_url: str, canvas_api_token: str, course_id: str) -> Optional[str]:
    """Upload an image to Canvas and return its URL."""
    try:
        # First, get upload parameters from Canvas
        files_url = f"{canvas_base_url}/api/v1/courses/{course_id}/files"

        with open(image_path, 'rb') as f:
            image_data = f.read()

        # Step 1: Tell Canvas we want to upload a file
        filename = os.path.basename(image_path)
        upload_params = {
            'name': filename,
            'size': len(image_data),
            'content_type': 'image/png' if filename.endswith('.png') else 'image/jpeg',
            'parent_folder_path': '/course_images'
        }

        headers = {'Authorization': f'Bearer {canvas_api_token}'}
        response = requests.post(files_url, data=upload_params, headers=headers)

        if response.status_code != 200:
            print(f"Failed to get upload URL for {filename}: {response.status_code}")
            print(response.text)
            return None

        upload_response = response.json()

        # Step 2: Upload the file to the URL provided by Canvas
        upload_url = upload_response['upload_url']
        upload_params = upload_response['upload_params']

        files = {'file': (filename, image_data, upload_params['content_type'])}

        # Remove 'file' from upload_params since we're passing it as files
        if 'file' in upload_params:
            del upload_params['file']

        upload_result = requests.post(upload_url, data=upload_params, files=files)

        if upload_result.status_code not in [200, 201]:
            print(f"Failed to upload {filename}: {upload_result.status_code}")
            print(upload_result.text)
            return None

        # Step 3: Get the file info from Canvas
        file_info = upload_result.json()
        if 'id' in file_info:
            file_id = file_info['id']
            # Return Canvas file URL
            return f"{canvas_base_url}/courses/{course_id}/files/{file_id}/preview"
        else:
            print(f"No file ID returned for {filename}")
            return None

    except Exception as e:
        print(f"Error uploading image {image_path}: {e}")
        return None

def apply_tint(rgb_hex: str, tint: float) -> str:
    """
    Apply a tint value to an RGB color.

    Tint values:
    - Positive tint: lightens the color (mixes with white)
    - Negative tint: darkens the color (mixes with black)
    - 0: no change

    Args:
        rgb_hex: Hex color string (e.g., 'FF0000' or '#FF0000')
        tint: Tint value between -1.0 and 1.0

    Returns:
        Hex color string with tint applied
    """
    # Remove '#' if present
    rgb_hex = rgb_hex.lstrip('#')

    # Convert hex to RGB
    r = int(rgb_hex[0:2], 16)
    g = int(rgb_hex[2:4], 16)
    b = int(rgb_hex[4:6], 16)

    # Apply tint
    if tint < 0:
        # Darken: mix with black
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    else:
        # Lighten: mix with white
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)

    # Clamp values to 0-255
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))

    return f'{r:02X}{g:02X}{b:02X}'

def get_theme_colors(workbook) -> Dict[int, str]:
    """
    Extract theme colors from workbook.

    Returns a dictionary mapping theme index to RGB hex color.
    """
    # Default Office theme colors (used if extraction fails)
    default_theme_colors = {
        0: 'FFFFFF',  # White (background 1)
        1: '000000',  # Black (text 1)
        2: 'E7E6E6',  # Light Gray (background 2)
        3: '44546A',  # Dark Blue Gray (text 2)
        4: '4472C4',  # Blue (accent 1)
        5: 'ED7D31',  # Orange (accent 2)
        6: 'A5A5A5',  # Gray (accent 3)
        7: 'FFC000',  # Gold (accent 4)
        8: '5B9BD5',  # Light Blue (accent 5)
        9: '70AD47',  # Green (accent 6)
    }

    try:
        # Try to extract colors from the loaded theme XML
        if hasattr(workbook, 'loaded_theme') and workbook.loaded_theme:
            theme_xml = workbook.loaded_theme.decode('utf-8')

            # Extract RGB colors from theme XML
            # Theme color indices:
            # 0=lt1 (light 1/background 1), 1=dk1 (dark 1/text 1)
            # 2=lt2 (light 2/background 2), 3=dk2 (dark 2/text 2)
            # 4=accent1, 5=accent2, 6=accent3, 7=accent4, 8=accent5, 9=accent6

            import xml.etree.ElementTree as ET
            root = ET.fromstring(theme_xml)

            # Define namespace
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

            # Find color scheme
            color_scheme = root.find('.//a:clrScheme', ns)
            if color_scheme:
                theme_colors = {}

                # Map color names to indices
                color_map = {
                    'a:lt1': 0,   # Light 1 (background 1)
                    'a:dk1': 1,   # Dark 1 (text 1)
                    'a:lt2': 2,   # Light 2 (background 2)
                    'a:dk2': 3,   # Dark 2 (text 2)
                    'a:accent1': 4,
                    'a:accent2': 5,
                    'a:accent3': 6,
                    'a:accent4': 7,
                    'a:accent5': 8,
                    'a:accent6': 9,
                }

                for color_name, index in color_map.items():
                    color_elem = color_scheme.find(color_name, ns)
                    if color_elem:
                        # Try to find srgbClr (direct RGB)
                        srgb = color_elem.find('.//a:srgbClr', ns)
                        if srgb is not None and 'val' in srgb.attrib:
                            theme_colors[index] = srgb.attrib['val']
                            continue

                        # Try to find sysClr (system color with lastClr attribute)
                        sysclr = color_elem.find('.//a:sysClr', ns)
                        if sysclr is not None and 'lastClr' in sysclr.attrib:
                            theme_colors[index] = sysclr.attrib['lastClr']
                            continue

                # If we extracted colors successfully, use them
                if len(theme_colors) >= 6:  # At least got the main colors
                    return theme_colors
    except Exception as e:
        # If extraction fails, fall back to defaults
        pass

    return default_theme_colors

def color_to_hex(color_obj, theme_colors: Dict[int, str]) -> Optional[str]:
    """
    Convert an openpyxl Color object to hex RGB string.

    Args:
        color_obj: openpyxl Color object
        theme_colors: Dictionary of theme colors

    Returns:
        Hex color string (e.g., 'FF0000') or None if no color
    """
    if not color_obj:
        return None

    try:
        # Direct RGB color
        if hasattr(color_obj, 'rgb') and color_obj.rgb:
            rgb_str = str(color_obj.rgb)
            # Check if it's an actual hex string (not a validation error message)
            if rgb_str and 'must be of type' not in rgb_str and rgb_str not in ['00000000', 'FFFFFFFF', '0', 'ffffffff']:
                # RGB values come as 8-char hex (AARRGGBB), we want RRGGBB
                if len(rgb_str) == 8:
                    return rgb_str[2:]
                elif len(rgb_str) == 6:
                    return rgb_str

        # Theme color with optional tint
        if hasattr(color_obj, 'theme'):
            theme_val = color_obj.theme
            # Validate that theme is actually an integer (not a validation error)
            if isinstance(theme_val, int) and theme_val in theme_colors:
                base_color = theme_colors[theme_val]

                # Apply tint if present
                if hasattr(color_obj, 'tint') and color_obj.tint:
                    return apply_tint(base_color, color_obj.tint)
                return base_color

        # Indexed color (legacy)
        if hasattr(color_obj, 'indexed'):
            idx_val = color_obj.indexed
            if isinstance(idx_val, int) and 0 <= idx_val < len(COLOR_INDEX):
                return COLOR_INDEX[idx_val].replace('#', '')

    except Exception:
        pass

    return None

def render_excel_sheet_to_html(excel_path: str, sheet_name: str = None) -> str:
    """
    Render an Excel sheet as an HTML table, preserving cell colors.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to render (if None, uses first sheet)

    Returns:
        HTML table string
    """
    if not OPENPYXL_AVAILABLE:
        return '<p><em>Note: Excel sheet rendering not available (openpyxl not installed)</em></p>'

    if not os.path.exists(excel_path):
        return f'<p><em>Error: Excel file not found: {excel_path}</em></p>'

    try:
        # Load workbook without data_only to preserve styling
        workbook = load_workbook(excel_path, data_only=False)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return f'<p><em>Error: Sheet "{sheet_name}" not found in {excel_path}</em></p>'
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # Get theme colors for the workbook
        theme_colors = get_theme_colors(workbook)

        # Build HTML table
        html_parts = ['<div style="overflow-x: auto; margin: 1.5em 0;">']
        html_parts.append('<table style="border-collapse: collapse; width: 100%; border: 1px solid #ddd;">')

        # Process each row
        first_row = True
        for row in sheet.iter_rows():
            # Skip completely empty rows
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                continue

            html_parts.append('<tr>')
            for cell in row:
                cell_value = str(cell.value) if cell.value is not None else ''

                # Extract cell background color using new helper
                bg_color = 'white'
                text_color = '#000000'

                # Extract background color if there's a fill
                if cell.fill and cell.fill.fill_type and cell.fill.fill_type != 'none':
                    if cell.fill.start_color:
                        bg_hex = color_to_hex(cell.fill.start_color, theme_colors)
                        if bg_hex:
                            bg_color = f'#{bg_hex}'

                # Extract text color if available
                if cell.font and cell.font.color:
                    text_hex = color_to_hex(cell.font.color, theme_colors)
                    # Only use the extracted color if it's not transparent/default
                    if text_hex and text_hex.upper() not in ['00000000', 'FFFFFFFF', 'FFFFFF', '000000']:
                        text_color = f'#{text_hex}'
                    # If text color matches background, use black for visibility
                    if text_color == bg_color:
                        text_color = '#000000'

                # Determine if bold
                is_bold = cell.font and cell.font.bold
                font_weight = 'bold' if is_bold else 'normal'

                if first_row:
                    # Header row styling - if no background color was set, use a nice header color
                    header_bg = bg_color if bg_color != 'white' else '#1976d2'
                    header_text = text_color if bg_color != 'white' else 'white'
                    header_weight = font_weight if font_weight == 'bold' else '600'

                    html_parts.append(
                        f'<th style="background-color: {header_bg}; color: {header_text}; '
                        f'padding: 12px 16px; text-align: left; font-weight: {header_weight}; '
                        f'border: 1px solid #ddd;">{cell_value}</th>'
                    )
                else:
                    # Regular cell styling
                    html_parts.append(
                        f'<td style="padding: 12px 16px; border: 1px solid #ddd; '
                        f'background-color: {bg_color}; color: {text_color}; '
                        f'font-weight: {font_weight};">{cell_value}</td>'
                    )
            html_parts.append('</tr>')
            first_row = False

        html_parts.append('</table>')
        html_parts.append('</div>')

        return ''.join(html_parts)

    except Exception as e:
        return f'<p><em>Error rendering Excel sheet: {str(e)}</em></p>'

def process_excel_macros(content: str, markdown_file_path: str) -> str:
    """
    Process Excel rendering macros in markdown content.
    Replaces {{ render_excel_sheet('./path/to/file.xlsx', 'SheetName') }} with rendered HTML tables.

    Args:
        content: Markdown content with Excel macros
        markdown_file_path: Path to the markdown file (for resolving relative paths)

    Returns:
        Content with Excel macros replaced by HTML tables
    """
    markdown_dir = os.path.dirname(os.path.abspath(markdown_file_path))

    # Pattern to match {{ render_excel_sheet('path', 'sheet') }}
    pattern = r'\{\{\s*render_excel_sheet\([\'"]([^\'"]+)[\'"](?:,\s*[\'"]([^\'"]+)[\'"])?\)\s*\}\}'

    def replace_excel_macro(match):
        excel_path = match.group(1)
        sheet_name = match.group(2) if match.group(2) else None

        # Resolve relative path
        if not os.path.isabs(excel_path):
            excel_path = os.path.join(markdown_dir, excel_path)

        excel_path = os.path.normpath(excel_path)

        print(f"  Rendering Excel sheet: {excel_path}" + (f" (sheet: {sheet_name})" if sheet_name else ""))

        return render_excel_sheet_to_html(excel_path, sheet_name)

    return re.sub(pattern, replace_excel_macro, content)

def process_images(content: str, markdown_file_path: str, canvas_base_url: str, canvas_api_token: str, course_id: str) -> str:
    """Process and upload images referenced in markdown content."""
    markdown_dir = os.path.dirname(os.path.abspath(markdown_file_path))

    # Find all image references in markdown
    image_pattern = r'!\[([^\]]*)\]\(([^)]+)\)'

    def replace_image(match):
        alt_text = match.group(1)
        image_path = match.group(2)

        # Skip if it's already a URL
        if image_path.startswith(('http://', 'https://')):
            return match.group(0)

        # Resolve relative path
        if not os.path.isabs(image_path):
            full_image_path = os.path.join(markdown_dir, image_path)
        else:
            full_image_path = image_path

        # Normalize the path
        full_image_path = os.path.normpath(full_image_path)

        if os.path.exists(full_image_path):
            print(f"Uploading image: {full_image_path}")
            canvas_url = upload_image_to_canvas(full_image_path, canvas_base_url, canvas_api_token, course_id)
            if canvas_url:
                print(f"✓ Uploaded: {os.path.basename(full_image_path)}")
                return f'<img src="{canvas_url}" alt="{alt_text}" />'
            else:
                print(f"✗ Failed to upload: {os.path.basename(full_image_path)}")
                return match.group(0)  # Keep original if upload failed
        else:
            print(f"Image not found: {full_image_path}")
            return match.group(0)  # Keep original if file not found

    return re.sub(image_pattern, replace_image, content)

def resolve_internal_links(content: str, all_pages_info: List[Dict], course_id: str, canvas_base_url: str) -> str:
    """Resolve internal HTML links to Canvas page URLs."""
    if not all_pages_info:
        return content

    # Create a mapping of potential markdown filenames to Canvas pages
    filename_to_page = {}
    for page_info in all_pages_info:
        page_url = page_info.get('url', '')
        page_title = page_info.get('title', '')

        # Try to extract potential original filename from title or URL
        # Method 1: Look for common patterns in the title
        if 'laplace' in page_title.lower() and 'mechanism' in page_title.lower():
            filename_to_page['lap-mech'] = page_info
        elif 'differential privacy' in page_title.lower() or 'pure' in page_title.lower():
            filename_to_page['dp-intro'] = page_info
        elif 'preliminaries' in page_title.lower():
            filename_to_page['dp-pre'] = page_info
        elif 'sensitivity' in page_title.lower():
            filename_to_page['sensitivity'] = page_info

        # Method 2: Also try exact endswith matching for other cases
        # Extract the last part of Canvas URL and see if it matches patterns
        url_parts = page_url.split('-')
        if len(url_parts) >= 2:
            # Try matching the main part without prefix and suffix numbers
            main_part = '-'.join(url_parts[1:-1]) if len(url_parts) > 2 else url_parts[0]
            if main_part not in filename_to_page:
                filename_to_page[main_part] = page_info

    def replace_link(match):
        link_url = match.group(1)
        link_text = match.group(2)

        # Skip external links
        if link_url.startswith(('http://', 'https://', 'mailto:')):
            return match.group(0)

        # Process relative markdown links (check if the base URL ends with .md)
        base_url = link_url.split('#')[0]  # Remove anchor for the check
        if base_url.endswith('.md'):
            # Extract anchor if present
            anchor = ""
            if '#' in link_url:
                link_url, anchor = link_url.split('#', 1)
                anchor = f"#{anchor}"

            # Normalize the path and convert to page slug format
            page_path = os.path.normpath(link_url)
            page_slug = os.path.splitext(os.path.basename(page_path))[0].replace('_', '-').lower()

            # First try our smart mapping
            if page_slug in filename_to_page:
                page_info = filename_to_page[page_slug]
                canvas_page_url = f"{canvas_base_url}/courses/{course_id}/pages/{page_info['url']}{anchor}"
                return f'<a href="{canvas_page_url}">{link_text}</a>'

            # Fallback: try exact endswith matching (original logic)
            for page_info in all_pages_info:
                if page_info.get('url', '').endswith(page_slug):
                    canvas_page_url = f"{canvas_base_url}/courses/{course_id}/pages/{page_info['url']}{anchor}"
                    return f'<a href="{canvas_page_url}">{link_text}</a>'

            # If no match found, keep original
            return match.group(0)

        return match.group(0)

    # Replace HTML-style links <a href="url">text</a>
    content = re.sub(r'<a href="([^"]+)">([^<]+)</a>', replace_link, content)
    return content

def robust_math_protection(content: str) -> str:
    """
    Robust math protection using character-by-character parsing.
    This avoids complex regex issues with multiline math content.
    """
    result = []
    i = 0
    length = len(content)

    while i < length:
        # Look for $$ (display math)
        if i < length - 1 and content[i:i+2] == '$$':
            # Find the closing $$
            j = i + 2
            while j < length - 1:
                if content[j:j+2] == '$$':
                    # Found closing $$
                    math_content = content[i+2:j]

                    # Process the math content
                    # Convert align to array FIRST
                    math_content = convert_align_to_array(math_content)
                    # Preserve line breaks
                    math_content = math_content.replace('\\\\', '§LATEXLINEBREAK§')
                    # Double all remaining backslashes to protect from markdown
                    math_content = math_content.replace('\\', '\\\\')
                    # Protect underscores, asterisks, and tildes from markdown processing
                    math_content = math_content.replace('_', '§UNDERSCORE§')
                    math_content = math_content.replace('*', '§ASTERISK§')
                    math_content = math_content.replace('~', '§TILDE§')

                    # Add protected math block
                    result.append(f'§MATH_DISPLAY_START§{math_content}§MATH_DISPLAY_END§')
                    i = j + 2
                    break
                j += 1
            else:
                # No closing $$ found, treat as regular character
                result.append(content[i])
                i += 1

        # Look for single $ (inline math) - but not if it's part of $$
        elif content[i] == '$' and (i == 0 or content[i-1] != '$') and (i == length-1 or content[i+1] != '$'):
            # Find the closing $
            j = i + 1
            while j < length and content[j] != '$':
                # Skip escaped dollars
                if content[j] == '\\' and j + 1 < length:
                    j += 2
                else:
                    j += 1

            if j < length:
                # Found closing $
                math_content = content[i+1:j]

                # Process inline math content
                # Double backslashes and protect characters
                math_content = math_content.replace('\\', '\\\\')
                math_content = math_content.replace('_', '§UNDERSCORE§')
                math_content = math_content.replace('*', '§ASTERISK§')
                math_content = math_content.replace('~', '§TILDE§')

                # Add protected inline math
                result.append(f'§MATH_INLINE_START§{math_content}§MATH_INLINE_END§')
                i = j + 1
            else:
                # No closing $ found, treat as regular character
                result.append(content[i])
                i += 1
        else:
            result.append(content[i])
            i += 1

    return ''.join(result)

def process_markdown_to_html(content: str, markdown_file_path: str, canvas_base_url: str,
                           canvas_api_token: str, course_id: str, all_pages_info: List[Dict] = None) -> str:
    """
    Complete processing pipeline to convert markdown to Canvas-compatible HTML.

    Args:
        content: Raw markdown content
        markdown_file_path: Path to the markdown file (for resolving relative images)
        canvas_base_url: Canvas base URL
        canvas_api_token: Canvas API token
        course_id: Canvas course ID
        all_pages_info: List of all Canvas pages for internal link resolution

    Returns:
        Processed HTML content ready for Canvas
    """
    # Step 0: Process Excel macros (before any other processing)
    content = process_excel_macros(content, markdown_file_path)

    # Step 1: Process LaTeX math environments and protect from markdown interference

    # Step 1a: Handle standalone align environments that need $$ wrapping
    # Use a simple approach that looks for \begin{align} not preceded by $$
    def wrap_standalone_align(content):
        """Wrap standalone align environments with $$ delimiters."""
        result = []
        i = 0
        length = len(content)

        while i < length:
            # Look for \begin{align} (single backslash) - 13 characters
            if content[i:i+13] == '\\begin{align}':
                # Check if it's preceded by $$ (looking backward for recent $$)
                j = i - 1
                found_dollar_signs = False
                # Look backward through whitespace and newlines to find $$
                while j >= 1:  # Need at least 2 characters for $$
                    if content[j-1:j+1] == '$$':
                        found_dollar_signs = True
                        break
                    # Only continue if we're seeing whitespace/newlines
                    if content[j] not in ' \t\n':
                        break
                    j -= 1

                if not found_dollar_signs:
                    # This is a standalone align, wrap it
                    # Find the matching \end{align}
                    align_start = i
                    k = i + 13  # Start after \begin{align}
                    align_depth = 1

                    while k < length and align_depth > 0:
                        if content[k:k+13] == '\\begin{align}':
                            align_depth += 1
                            k += 13
                        elif content[k:k+11] == '\\end{align}':
                            align_depth -= 1
                            if align_depth == 0:
                                k += 11
                                break
                            else:
                                k += 11
                        else:
                            k += 1

                    if align_depth == 0:
                        # Found complete align block, wrap it
                        align_content = content[align_start:k]
                        result.append(f'$${align_content}$$')
                        i = k
                    else:
                        # Incomplete align block, leave as is
                        result.append(content[i])
                        i += 1
                else:
                    result.append(content[i])
                    i += 1
            else:
                result.append(content[i])
                i += 1

        return ''.join(result)

    content = wrap_standalone_align(content)

    # Step 1b: Use robust character-by-character math protection
    content = robust_math_protection(content)

    # Step 2: Process images (upload to Canvas)
    content = process_images(content, markdown_file_path, canvas_base_url, canvas_api_token, course_id)

    # Step 3: Convert markdown to HTML (math is already protected)
    md = markdown.Markdown(extensions=[
        'extra',           # Includes tables, fenced code blocks, etc.
        'codehilite',      # Syntax highlighting with Pygments
        'toc',             # Table of contents
        'sane_lists',      # Better list handling
        'admonition',      # Note/warning blocks
        'pymdownx.details',      # Support collapsible notes
        'pymdownx.superfences',  # (requirement of pymdownx.details)
        'attr_list',       # Attribute lists
        'def_list',        # Definition lists
        'footnotes',       # Footnotes
        'meta'             # Metadata
    ], extension_configs={
        'codehilite': {
            'css_class': 'codehilite',
            'guess_lang': True,
            'linenums': False,
            'use_pygments': True
        }
    })

    html_content = md.convert(content)

    # Step 4: Restore protected math content
    html_content = restore_protected_math_content(html_content)

    # Step 6: Add Pygments inline styles for Canvas compatibility
    html_content = add_pygments_inline_styles(html_content)

    # Step 7: Style code blocks and admonitions
    html_content = style_code_blocks(html_content)
    html_content = style_admonitions(html_content)

    # Step 8: Resolve internal links (if page info provided)
    if all_pages_info:
        html_content = resolve_internal_links(html_content, all_pages_info, course_id, canvas_base_url)

    return html_content

def restore_protected_math_content(content: str) -> str:
    """Restore protected math content with Canvas-compatible delimiters."""
    # Restore math delimiters - Canvas uses \(...\) for inline and $$...$$ for display
    content = content.replace('§MATH_INLINE_START§', r'\(')
    content = content.replace('§MATH_INLINE_END§', r'\)')
    content = content.replace('§MATH_DISPLAY_START§', '$$')
    content = content.replace('§MATH_DISPLAY_END§', '$$')

    # Fix double-escaped backslashes and restore protected characters in math
    def fix_math_backslashes(match):
        math_content = match.group(0)
        # Undo the backslash doubling we did during protection
        # We doubled all \ to \\ to protect them through markdown
        # Now we need to convert them back: \\\\ -> \\
        math_content = math_content.replace('\\\\', '\\')
        # Restore special LaTeX line breaks as double backslashes for MathJax
        # Use regex to handle any spacing variations around the token
        math_content = re.sub(r'\s*§LATEXLINEBREAK§\s*', r' \\\\ ', math_content)
        # Restore asterisks, underscores, and tildes that we protected
        math_content = math_content.replace('§ASTERISK§', '*')
        math_content = math_content.replace('§UNDERSCORE§', '_')
        math_content = math_content.replace('§TILDE§', '~')
        return math_content

    # Fix in both display ($$...$$) and inline (\(...\)) math
    content = re.sub(r'\$\$.*?\$\$', fix_math_backslashes, content, flags=re.DOTALL)
    content = re.sub(r'\\\(.*?\\\)', fix_math_backslashes, content, flags=re.DOTALL)

    return content

def add_pygments_inline_styles(html_content: str) -> str:
    """Add inline styles for Pygments syntax highlighting (GitHub light theme)"""
    pygments_styles = {
        'class="c"': 'class="c" style="color: #6a737d; font-style: italic;"',
        'class="k"': 'class="k" style="color: #d73a49; font-weight: 600;"',
        'class="n"': 'class="n" style="color: #24292e;"',
        'class="o"': 'class="o" style="color: #d73a49;"',
        'class="p"': 'class="p" style="color: #24292e;"',
        'class="cm"': 'class="cm" style="color: #6a737d; font-style: italic;"',
        'class="c1"': 'class="c1" style="color: #6a737d; font-style: italic;"',
        'class="kc"': 'class="kc" style="color: #005cc5;"',
        'class="kd"': 'class="kd" style="color: #d73a49; font-weight: 600;"',
        'class="kn"': 'class="kn" style="color: #d73a49; font-weight: 600;"',
        'class="kp"': 'class="kp" style="color: #d73a49; font-weight: 600;"',
        'class="kr"': 'class="kr" style="color: #d73a49; font-weight: 600;"',
        'class="kt"': 'class="kt" style="color: #005cc5; font-weight: 600;"',
        'class="m"': 'class="m" style="color: #005cc5;"',
        'class="s"': 'class="s" style="color: #032f62;"',
        'class="na"': 'class="na" style="color: #22863a;"',
        'class="nb"': 'class="nb" style="color: #005cc5;"',
        'class="nc"': 'class="nc" style="color: #6f42c1; font-weight: 600;"',
        'class="no"': 'class="no" style="color: #005cc5;"',
        'class="nd"': 'class="nd" style="color: #6f42c1;"',
        'class="nf"': 'class="nf" style="color: #6f42c1; font-weight: 600;"',
        'class="nn"': 'class="nn" style="color: #24292e;"',
        'class="nt"': 'class="nt" style="color: #22863a;"',
        'class="nv"': 'class="nv" style="color: #e36209;"',
        'class="ow"': 'class="ow" style="color: #d73a49; font-weight: 600;"',
        'class="w"': 'class="w" style="color: #24292e;"',
        'class="mf"': 'class="mf" style="color: #005cc5;"',
        'class="mh"': 'class="mh" style="color: #005cc5;"',
        'class="mi"': 'class="mi" style="color: #005cc5;"',
        'class="mo"': 'class="mo" style="color: #005cc5;"',
        'class="sb"': 'class="sb" style="color: #032f62;"',
        'class="sc"': 'class="sc" style="color: #032f62;"',
        'class="sd"': 'class="sd" style="color: #032f62;"',
        'class="s2"': 'class="s2" style="color: #032f62;"',
        'class="se"': 'class="se" style="color: #005cc5;"',
        'class="sh"': 'class="sh" style="color: #032f62;"',
        'class="si"': 'class="si" style="color: #005cc5;"',
        'class="sx"': 'class="sx" style="color: #032f62;"',
        'class="sr"': 'class="sr" style="color: #032f62;"',
        'class="s1"': 'class="s1" style="color: #032f62;"',
        'class="ss"': 'class="ss" style="color: #032f62;"',
    }

    for old, new in pygments_styles.items():
        html_content = html_content.replace(old, new)

    return html_content

def style_code_blocks(html_content: str) -> str:
    """Add inline styles to code blocks for Canvas compatibility"""
    # Style code block containers
    html_content = html_content.replace(
        '<div class="codehilite">',
        '<div class="codehilite" style="background-color: #f6f8fa; padding: 16px; border-radius: 6px; margin: 1.5em 0; border: 1px solid #e1e4e8;">'
    )

    # Style pre elements
    html_content = html_content.replace(
        '<pre>',
        '<pre style="background-color: #f6f8fa; margin: 0; overflow-x: auto; color: #24292e; line-height: 1.5;">'
    )

    # Style inline code elements
    html_content = html_content.replace(
        '<code>',
        '<code style="background-color: transparent; font-family: \'SFMono-Regular\', \'Consolas\', \'Liberation Mono\', \'Menlo\', monospace; font-size: 0.9em; color: #24292e;">'
    )

    return html_content

def style_admonitions(html_content: str) -> str:
    """Add inline styles for admonitions (Note, Warning, etc.)"""
    admonition_styles = {
        'note': 'background-color: #e7f2fa; border-left: 4px solid #2196F3; color: #014361;',
        'warning': 'background-color: #fff4e5; border-left: 4px solid #ff9800; color: #663c00;',
        'important': 'background-color: #ffe5e5; border-left: 4px solid #f44336; color: #5f2120;',
        'tip': 'background-color: #e8f5e9; border-left: 4px solid #4caf50; color: #1b5e20;',
        'danger': 'background-color: #ffebee; border-left: 4px solid #f44336; color: #5f2120;',
        'info': 'background-color: #e1f5fe; border-left: 4px solid #03a9f4; color: #01579b;',
    }

    for admonition_type, style in admonition_styles.items():
        # Style the admonition container
        html_content = html_content.replace(
            f'<div class="admonition {admonition_type}">',
            f'<div class="admonition {admonition_type}" style="padding: 15px 20px; margin: 1.5em 0; border-radius: 4px; {style}">'
        )
        # Style the title
        html_content = html_content.replace(
            f'<p class="admonition-title">',
            f'<p class="admonition-title" style="font-weight: 600; margin: 0 0 10px 0; font-size: 1.1em;">'
        )

    return html_content
